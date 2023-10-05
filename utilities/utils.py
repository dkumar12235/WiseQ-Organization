import openpyxl
from openpyxl import Workbook, load_workbook

class TestDataLoader:
    @staticmethod
    def get_data(sheet_name):
        workbook = openpyxl.load_workbook("D:\\WiseQAutomation\\WiseqOrganization\\testdata\\orgtestdata.xlsx")
        sheet = workbook[sheet_name]

        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)

        return data
